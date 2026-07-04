import io
import csv
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# Mock default data for corporate reporting if DB is unpopulated
DEFAULT_REPORT_DATA = {
    "organization_score": 78,
    "total_users_checked": 1420,
    "distribution": {
        "Critical": 45,
        "Weak": 182,
        "Moderate": 431,
        "Strong": 512,
        "Excellent": 250
    },
    "compliance": {
        "NIST_800_63B": "85% Compliant",
        "OWASP_Top_10": "92% Compliant",
        "CIS_Controls": "76% Compliant",
    },
    "authentication_methods": {
        "Password Only": 412,
        "Password + SMS MFA": 612,
        "Password + TOTP MFA": 296,
        "Passkeys / FIDO2": 100
    },
    "risks": [
        "45 users have passwords in active breach datasets (breached password checker match).",
        "182 users have weak passwords containing sequential numbers or dictionary words.",
        "412 enterprise accounts lack multi-factor authentication enforcement."
    ],
    "recommendations": [
        "Enforce a minimum length of 12 characters across all divisions.",
        "Integrate automated sign-up verification against the HaveIBeenPwned API.",
        "Migrate legacy authentication instances to support WebAuthn / Passkeys."
    ]
}

def generate_csv_report() -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(["Enterprise Password Security Audit Report"])
    writer.writerow(["Generated At", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")])
    writer.writerow([])
    
    writer.writerow(["Section", "Metric", "Value"])
    writer.writerow(["Summary", "Overall Security Score", DEFAULT_REPORT_DATA["organization_score"]])
    writer.writerow(["Summary", "Total Users Audited", DEFAULT_REPORT_DATA["total_users_checked"]])
    
    writer.writerow([])
    writer.writerow(["Password Strength Distribution"])
    for cat, val in DEFAULT_REPORT_DATA["distribution"].items():
        writer.writerow(["Distribution", cat, val])
        
    writer.writerow([])
    writer.writerow(["Standards Compliance Status"])
    for standard, status in DEFAULT_REPORT_DATA["compliance"].items():
        writer.writerow(["Compliance", standard, status])
        
    writer.writerow([])
    writer.writerow(["Authentication Methods Summary"])
    for method, count in DEFAULT_REPORT_DATA["authentication_methods"].items():
        writer.writerow(["Authentication", method, count])
        
    return output.getvalue().encode("utf-8")

def generate_json_report() -> bytes:
    report = {
        "report_type": "Enterprise Password Security & Authentication Intelligence Audit",
        "generated_at": datetime.utcnow().isoformat(),
        "summary": {
            "overall_security_score": DEFAULT_REPORT_DATA["organization_score"],
            "total_users_audited": DEFAULT_REPORT_DATA["total_users_checked"]
        },
        "distribution": DEFAULT_REPORT_DATA["distribution"],
        "compliance": DEFAULT_REPORT_DATA["compliance"],
        "authentication_methods": DEFAULT_REPORT_DATA["authentication_methods"],
        "identified_risks": DEFAULT_REPORT_DATA["risks"],
        "action_recommendations": DEFAULT_REPORT_DATA["recommendations"]
    }
    return json.dumps(report, indent=4).encode("utf-8")

def generate_excel_report() -> bytes:
    wb = Workbook()
    
    # Sheet 1: Executive Summary
    ws1 = wb.active
    ws1.title = "Executive Summary"
    
    # Title Block
    ws1["A1"] = "Enterprise Password Security Audit Report"
    ws1["A1"].font = Font(name="Arial", size=16, bold=True, color="1F497D")
    ws1["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
    ws1["A2"].font = Font(name="Arial", size=10, italic=True)
    
    # Headers
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    
    ws1["A4"] = "Overall Metric"
    ws1["B4"] = "Value"
    ws1["A4"].fill = header_fill
    ws1["B4"].fill = header_fill
    ws1["A4"].font = header_font
    ws1["B4"].font = header_font
    
    ws1["A5"] = "Overall Security Score"
    ws1["B5"] = f"{DEFAULT_REPORT_DATA['organization_score']}/100"
    ws1["A6"] = "Total Users Audited"
    ws1["B6"] = DEFAULT_REPORT_DATA["total_users_checked"]
    
    # Distribution Block
    ws1["A9"] = "Password Strength Category"
    ws1["B9"] = "User Count"
    ws1["A9"].fill = header_fill
    ws1["B9"].fill = header_fill
    ws1["A9"].font = header_font
    ws1["B9"].font = header_font
    
    row_idx = 10
    for cat, count in DEFAULT_REPORT_DATA["distribution"].items():
        ws1.cell(row=row_idx, column=1, value=cat)
        ws1.cell(row=row_idx, column=2, value=count)
        row_idx += 1
        
    # Sheet 2: Compliance & Risks
    ws2 = wb.create_sheet(title="Compliance & Action Items")
    ws2["A1"] = "Compliance Standards & Recommendations"
    ws2["A1"].font = Font(name="Arial", size=14, bold=True, color="1F497D")
    
    ws2["A3"] = "Compliance Framework"
    ws2["B3"] = "Compliance Level"
    ws2["A3"].fill = header_fill
    ws2["B3"].fill = header_fill
    ws2["A3"].font = header_font
    ws2["B3"].font = header_font
    
    ws2["A4"] = "NIST SP 800-63B Guidelines"
    ws2["B4"] = DEFAULT_REPORT_DATA["compliance"]["NIST_800_63B"]
    ws2["A5"] = "OWASP Top 10 Authentication"
    ws2["B5"] = DEFAULT_REPORT_DATA["compliance"]["OWASP_Top_10"]
    ws2["A6"] = "CIS Controls (Credential Standard)"
    ws2["B6"] = DEFAULT_REPORT_DATA["compliance"]["CIS_Controls"]
    
    # Recommendations
    ws2["A9"] = "Critical Recommendations"
    ws2["A9"].font = Font(name="Arial", size=12, bold=True, color="C00000")
    
    rec_row = 10
    for rec in DEFAULT_REPORT_DATA["recommendations"]:
        ws2.cell(row=rec_row, column=1, value=f"- {rec}")
        rec_row += 1
        
    # Auto-fit columns
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Save workbook to memory buffer
    file_stream = io.BytesIO()
    wb.save(file_stream)
    return file_stream.getvalue()

def generate_pdf_report() -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
    story = []
    
    styles = getSampleStyleSheet()
    
    # Create Premium custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Title'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.HexColor("#1e293b"),
        spaceAfter=15,
        alignment=0
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=10,
        textColor=colors.HexColor("#64748b"),
        spaceAfter=30
    )
    
    h1_style = ParagraphStyle(
        'H1Style',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=colors.HexColor("#0f172a"),
        spaceBefore=15,
        spaceAfter=10
    )
    
    body_style = ParagraphStyle(
        'BodyStyle',
        parent=styles['BodyText'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor("#334155"),
        spaceAfter=8,
        leading=14
    )
    
    bold_body = ParagraphStyle(
        'BoldBody',
        parent=body_style,
        fontName='Helvetica-Bold'
    )
    
    # --- Title Page Content ---
    story.append(Paragraph("Enterprise Password Security & Authentication Intelligence Report", title_style))
    story.append(Paragraph(f"Platform: Password Security Intelligence Suite | Generated: {datetime.utcnow().strftime('%B %d, %Y - %H:%M UTC')}", subtitle_style))
    
    # --- Section 1: Executive Summary ---
    story.append(Paragraph("1. Executive Summary", h1_style))
    story.append(Paragraph(
        f"This report evaluates the password policies, credential parameters, and security exposures across "
        f"{DEFAULT_REPORT_DATA['total_users_checked']} active corporate directory accounts. "
        f"The platform assessed overall password strength distribution, hashing standards compliance, "
        f"and multi-factor authentication (MFA) enforcement metrics.",
        body_style
    ))
    story.append(Spacer(1, 10))
    
    # Metric Summary Box
    summary_data = [
        [Paragraph("Overall Security Rating", bold_body), Paragraph(f"{DEFAULT_REPORT_DATA['organization_score']}/100", bold_body)],
        [Paragraph("Total Directory Users Checked", body_style), Paragraph(str(DEFAULT_REPORT_DATA['total_users_checked']), body_style)]
    ]
    t_summary = Table(summary_data, colWidths=[250, 150])
    t_summary.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#f8fafc")),
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor("#cbd5e1")),
        ('PADDING', (0,0), (-1,-1), 8),
        ('ALIGN', (1,0), (1,-1), 'RIGHT')
    ]))
    story.append(t_summary)
    story.append(Spacer(1, 20))
    
    # --- Section 2: Password Distribution ---
    story.append(Paragraph("2. Password Strength Distribution", h1_style))
    story.append(Paragraph("The checked passwords fall into the following strength categories based on length, complexity, and entropy calculations:", body_style))
    story.append(Spacer(1, 5))
    
    dist_data = [[Paragraph("Strength Category", bold_body), Paragraph("User Accounts Count", bold_body)]]
    for cat, val in DEFAULT_REPORT_DATA["distribution"].items():
        dist_data.append([Paragraph(cat, body_style), Paragraph(str(val), body_style)])
        
    t_dist = Table(dist_data, colWidths=[200, 200])
    t_dist.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#0f172a")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor("#e2e8f0")),
        ('PADDING', (0,0), (-1,-1), 6),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#f8fafc")]),
        ('ALIGN', (1,0), (1,-1), 'RIGHT')
    ]))
    story.append(t_dist)
    story.append(Spacer(1, 20))
    
    # --- Section 3: Compliance ---
    story.append(Paragraph("3. Regulatory Compliance Findings", h1_style))
    comp_data = [[Paragraph("Security Standard", bold_body), Paragraph("Status", bold_body)]]
    for standard, status in DEFAULT_REPORT_DATA["compliance"].items():
        comp_data.append([Paragraph(standard.replace("_", " "), body_style), Paragraph(status, bold_body)])
        
    t_comp = Table(comp_data, colWidths=[250, 150])
    t_comp.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor("#e2e8f0")),
        ('PADDING', (0,0), (-1,-1), 6),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#f1f5f9"))
    ]))
    story.append(t_comp)
    story.append(Spacer(1, 20))
    
    # --- Section 4: Risks and Recommendations ---
    story.append(Paragraph("4. Key Exposure Risks & Action Recommendations", h1_style))
    story.append(Paragraph("Identified Risks:", bold_body))
    for risk in DEFAULT_REPORT_DATA["risks"]:
        story.append(Paragraph(f"• {risk}", body_style))
        
    story.append(Spacer(1, 10))
    story.append(Paragraph("Remediation Action Items:", bold_body))
    for rec in DEFAULT_REPORT_DATA["recommendations"]:
        story.append(Paragraph(f"✔ {rec}", body_style))
        
    # Build Document
    doc.build(story)
    return buffer.getvalue()
